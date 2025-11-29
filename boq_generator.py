
# boq_generator_upgraded.py
"""
Upgraded BoQ generator (full) — fixes canonical mapping and quantity population.

Features:
- Stronger mapping from parsed Element/Description -> BESMM4 canonical items
- Diagnostic mode to print mapping decisions
- Preserves geometry fields so geometry_rules can compute quantities
- Uses provided Excel template and writes quantities, rates, amounts
- Audit sheet with justifications
Requires:
    pandas, openpyxl, ai_pricing, geometry_rules.py
"""

import os
from collections import defaultdict, OrderedDict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from ai_pricing import get_rate_from_library
from geometry_rules import RULE_REGISTRY

# ---------- Config ----------
DEFAULT_RATE = 0.0
TEMPLATE_PATH = os.path.expanduser("~/qs_ai_project/templates/besmm4_full_boq.xlsx")
DEFAULTS = {
    "room_height": 3.0,
    "slab_depth": 0.15,
}
DIAGNOSTIC = False  # set True to print mapping diagnostics

# User-provided mapping from parser element -> BESMM4 canonical (human-friendly)
ELEMENT_TO_BESMM4 = {
    "floor finish": "Tiling",           # maps Floor Finish -> Tiling (BESMM4 canonical)
    "wall finish": "Plastering",        # map wall finish -> Plastering (or Painting handled separately)
    "skirting": "Skirting",
    "ceiling finish": "Ceiling Finish",
    "windows": "Windows",
    "doors": "Doors",
    "excavation": "Excavation",
    # add more as needed
}

# Official high-level BESMM4 sections (1.0..8.0)
BESMM4_SECTIONS = OrderedDict([
    ("1.0", "Preliminaries"),
    ("2.0", "Substructure"),
    ("3.0", "Superstructure"),
    ("4.0", "Finishes"),
    ("5.0", "Fittings & Fixtures"),
    ("6.0", "Mechanical & Electrical Works"),
    ("7.0", "External Works"),
    ("8.0", "Provisional & Prime Cost Sums"),
])

# Hierarchy (kept minimal / same as original)
BESMM4_HIERARCHY = {
    "2.0": [
        ("2.1", "Excavating & Filling", [
            ("2.1.1", "Bulk excavation", "m3", "Excavation"),
            ("2.1.2", "Foundation excavation", "m3", "Excavation"),
            ("2.1.3", "Disposal off-site of excavated material", "m3", "Excavation"),
            ("2.1.4", "Filling obtained from excavated material", "m3", "Earthworks"),
        ]),
        ("2.2", "Concrete Work", [
            ("2.2.1", "Blinding concrete", "m3", "Concrete"),
            ("2.2.2", "Reinforced concrete in foundations", "m3", "Reinforced Concrete"),
            ("2.2.3", "Reinforced concrete ground floor slab", "m2", "Ground Floor Slab"),
        ]),
        ("2.3", "Damp Proofing & Membranes", [
            ("2.3.1", "Damp proof membrane", "m2", "Damp Proof Membrane"),
        ]),
    ],
    "3.0": [
        ("3.1", "Masonry / Blockwork", [
            ("3.1.1", "Blockwork 225mm thick in cement mortar", "m2", "Blockwork"),
            ("3.1.2", "Partition walls", "m2", "Partition Walls"),
        ]),
        ("3.2", "Structural Concrete & Steel", [
            ("3.2.1", "Reinforced concrete columns", "m3", "Reinforced Concrete"),
            ("3.2.2", "Structural steelwork", "t", "Steelwork"),
        ]),
        ("3.3", "Openings", [
            ("3.3.1", "Doors (supply & fix)", "No.", "Doors"),
            ("3.3.2", "Windows (supply & fix)", "No.", "Windows"),
        ]),
    ],
    "4.0": [
        ("4.1", "Internal Finishes", [
            ("4.1.1", "Internal plastering 12mm", "m2", "Plastering"),
            ("4.1.2", "Floor tiling (ceramic)", "m2", "Tiling"),
            ("4.1.3", "Skirting (supply & fix)", "m", "Skirting"),
            ("4.1.4", "Ceiling finishes (gypsum/POP/PVC)", "m2", "Ceiling Finish"),
            ("4.1.5", "Painting to walls (2 coats)", "m2", "Painting"),
        ]),
    ],
    "5.0": [
        ("5.1", "Fittings & Fixtures", [
            ("5.1.1", "Sanitary fittings (WC, basin, etc.)", "No.", "Sanitary Fittings"),
            ("5.1.2", "Kitchen fittings (complete)", "No.", "Kitchen Fittings"),
            ("5.1.3", "Wardrobes (supply & fix)", "No.", "Wardrobes"),
        ]),
    ],
    "6.0": [
        ("6.1", "Mechanical Services", [
            ("6.1.1", "Plumbing fittings (points)", "No.", "Plumbing"),
            ("6.1.2", "HVAC system (per system)", "No.", "HVAC"),
        ]),
        ("6.2", "Electrical Services", [
            ("6.2.1", "Lighting fittings", "No.", "Lighting"),
            ("6.2.2", "Power points", "No.", "Small Power"),
            ("6.2.3", "CCTV", "No.", "CCTV"),
        ]),
    ],
    "7.0": [
        ("7.1", "External Works", [
            ("7.1.1", "Paving", "m2", "Paving"),
            ("7.1.2", "Driveways", "m2", "Driveways"),
            ("7.1.3", "Landscaping", "m2", "Landscaping"),
        ]),
    ],
    "8.0": [
        ("8.1", "Provisional Sums", [
            ("8.1.1", "Specialized works (PC sum)", "item", "Specialized Works"),
            ("8.1.2", "Contingencies", "item", "Contingencies"),
        ]),
    ],
}

# Build flat hierarchy map
def _flatten_hierarchy():
    flat = {}
    for sec_code, groups in BESMM4_HIERARCHY.items():
        sec_name = BESMM4_SECTIONS.get(sec_code, sec_code)
        for g_code, g_name, items in groups:
            for item_code, item_desc, unit, canonical in items:
                flat_key = (canonical or item_desc).lower()
                flat[flat_key] = {
                    "section_code": sec_code,
                    "section_name": sec_name,
                    "group_code": g_code,
                    "group_name": g_name,
                    "item_code": item_code,
                    "item_desc": item_desc,
                    "unit": unit,
                    "canonical": canonical or item_desc
                }
    return flat

FLAT_HIERARCHY = _flatten_hierarchy()

# Lightweight keyword mapping
ELEMENT_KEYWORDS = {
    "excavation": "Excavation",
    "earthwork": "Earthworks",
    "foundation": "Foundations",
    "ground floor slab": "Ground Floor Slab",
    "slab": "Ground Floor Slab",
    "blockwork": "Blockwork",
    "block": "Blockwork",
    "masonry": "Blockwork",
    "partition": "Partition Walls",
    "plaster": "Plastering",
    "plastering": "Plastering",
    "tiling": "Tiling",
    "tile": "Tiling",
    "painting": "Painting",
    "paint": "Painting",
    "ceiling": "Ceiling Finish",
    "skirting": "Skirting",
    "skirt": "Skirting",
    "door": "Doors",
    "window": "Windows",
    "sanitary": "Sanitary Fittings",
    "kitchen": "Kitchen Fittings",
    "plumbing": "Plumbing",
    "hvac": "HVAC",
    "lighting": "Lighting",
    "power": "Small Power",
    "cctv": "CCTV",
    "paving": "Paving",
    "driveway": "Driveways",
    "landscap": "Landscaping",
    "special": "Specialized Works",
    "conting": "Contingencies",
    "damp proof": "Damp Proof Membrane",
    "dpm": "Damp Proof Membrane",
    "reinforced concrete": "Reinforced Concrete",
    "concrete": "Concrete",
    "steel": "Steelwork",
}

# ---------------- Utilities ----------------

def _normalize_text(s):
    if s is None:
        return ""
    return str(s).strip().lower()

def _canonical_element_from_text(text: str):
    """Try to find a canonical element name from arbitrary text using expanded keywords."""
    if text is None:
        return None
    t = _normalize_text(text)
    # direct keyword match
    for k, canonical in ELEMENT_KEYWORDS.items():
        if k in t:
            return canonical
    return None

# ---------------- Aggregation & Mapping ----------------

def aggregate_parsed_entries(parsed_entries):
    """
    Aggregate parsed entries but preserve geometry and raw entries for rule handlers.
    Also performs a mapping from parsed Element -> BESMM4 canonical via ELEMENT_TO_BESMM4 and ELEMENT_KEYWORDS.
    Returns: dict keyed by canonical.lower() -> payload
    """
    agg = {}
    for e in parsed_entries:
        element_raw = e.get("Element") or ""
        description_raw = e.get("Description") or ""
        element_norm = _normalize_text(element_raw)
        description_norm = _normalize_text(description_raw)
        unit = (e.get("Unit") or "item") or "item"

        # try user mapping first
        mapped = None
        if element_norm in ELEMENT_TO_BESMM4:
            mapped = ELEMENT_TO_BESMM4[element_norm]
        else:
            # try keyword-based canonical from element or description
            mapped = _canonical_element_from_text(element_raw) or _canonical_element_from_text(description_raw)

        # if still none, fallback to element text (first word) or description
        if not mapped:
            if element_raw:
                mapped = element_raw
            else:
                mapped = description_raw

        canonical = mapped or ""
        key = canonical.lower()

        # preserved data for geometry
        preserved = {
            "Room": e.get("Room"),
            "Element": element_raw,
            "Description": description_raw,
            "Unit": unit,
            "Quantity": e.get("Quantity", e.get("Qty", None)),
            "length": e.get("length") or e.get("Length"),
            "width": e.get("width") or e.get("Width"),
            "height": e.get("height") or e.get("Height"),
            "area": e.get("area") or e.get("Area"),
            "thickness": e.get("thickness") or e.get("depth"),
            "openings": e.get("openings") or e.get("Openings") or e.get("openings_area"),
            "raw": e
        }

        if key not in agg:
            agg[key] = {"units": defaultdict(float), "descriptions": set(), "entries": []}
        # numeric aggregation for fallback
        try:
            q = e.get("Quantity", e.get("Qty", 0)) or 0
            qn = float(str(q).replace(",", "")) if q is not None else 0.0
        except Exception:
            qn = 0.0
        agg[key]["units"][unit] += qn
        if description_raw:
            agg[key]["descriptions"].add(description_raw)
        agg[key]["entries"].append(preserved)

        if DIAGNOSTIC:
            print("AGGREGATE:", element_raw, "|", description_raw, "=> canonical:", canonical)

    return agg

# ---------------- Geometry-based computation ----------------

def compute_quantities_from_geometry(agg, defaults=None):
    """
    Compute quantities using handlers based on canonical keys present in agg.
    Returns dict canonical.lower() -> {quantity, unit, justification}
    """
    if defaults is None:
        defaults = DEFAULTS
    computed = {}
    for canonical_key, payload in agg.items():
        # handler lookup: try direct FLAT_HIERARCHY canonical match or lowercased keys in RULE_REGISTRY
        handler_key = canonical_key  # handlers expect keys like 'plastering', 'tiling', etc.
        handler = RULE_REGISTRY.get(handler_key) or RULE_REGISTRY.get(canonical_key.lower()) or RULE_REGISTRY.get(_normalize_text(canonical_key))
        if handler is None:
            # try elemental keyword mapping
            handler = RULE_REGISTRY.get(_canonical_element_from_text(canonical_key))
        if handler is None:
            # fallback: use aggregated numeric sum across units
            units = payload.get("units", {})
            qty = 0.0
            if units:
                try:
                    qty = max(units.values())
                except Exception:
                    qty = sum(units.values())
            computed[canonical_key] = {
                "quantity": round(float(qty), 4),
                "unit": next(iter(units.keys())) if units else None,
                "justification": "Fallback to parsed numeric quantities"
            }
            if DIAGNOSTIC:
                print("NO HANDLER for:", canonical_key, "-> fallback qty:", qty)
        else:
            try:
                res = handler(payload.get("entries", []), defaults)
                q = float(res.get("quantity") or 0.0)
                computed[canonical_key] = {
                    "quantity": round(q, 4),
                    "unit": res.get("unit"),
                    "justification": res.get("justification", "")
                }
                if DIAGNOSTIC:
                    print("HANDLED:", canonical_key, "=>", computed[canonical_key])
            except Exception as ex:
                units = payload.get("units", {})
                qty = sum(units.values()) if units else 0.0
                computed[canonical_key] = {
                    "quantity": round(float(qty), 4),
                    "unit": next(iter(units.keys())) if units else None,
                    "justification": f"Handler error; fallback to parsed qty ({ex})"
                }
                if DIAGNOSTIC:
                    print("HANDLER ERROR for:", canonical_key, ex)
    return computed

# ---------------- Template population ----------------

def _build_full_item_list(include_empty=True):
    items = []
    for sec_code, groups in BESMM4_HIERARCHY.items():
        sec_name = BESMM4_SECTIONS.get(sec_code, sec_code)
        for g_code, g_name, itms in groups:
            for item_code, item_desc, unit, canonical in itms:
                items.append({
                    "SectionCode": sec_code,
                    "SectionName": sec_name,
                    "GroupCode": g_code,
                    "GroupName": g_name,
                    "ItemCode": item_code,
                    "Description": item_desc,
                    "Unit": unit,
                    "Canonical": (canonical or item_desc),
                    "Quantity": 0.0,
                    "Rate": 0.0,
                    "Amount": 0.0,
                    "Justification": ""
                })
    return items

def populate_besmm4_from_parsed(parsed_entries, location="Kaduna", include_empty=True):
    """
    Map parsed entries into BESMM4 master items using computed geometry quantities.
    """
    if DIAGNOSTIC:
        print("Starting populate_besmm4_from_parsed with", len(parsed_entries), "entries")

    agg = aggregate_parsed_entries(parsed_entries)
    computed = compute_quantities_from_geometry(agg, defaults=DEFAULTS)
    items = _build_full_item_list(include_empty=include_empty)

    # Create reverse map: canonical.lower() -> item record(s)
    canonical_to_items = defaultdict(list)
    for it in items:
        canonical_to_items[(it["Canonical"] or it["Description"]).lower()].append(it)

    # Map computed quantities into template items
    for canonical_key, res in computed.items():
        qty = res.get("quantity", 0.0)
        just = res.get("justification", "")
        # Try direct match in canonical_to_items
        if canonical_key in canonical_to_items:
            for it in canonical_to_items[canonical_key]:
                it["Quantity"] = float(round(qty, 4))
                it["Justification"] = just
        else:
            # fuzzy match: try substring or keyword match against keys
            matched = False
            for can_k, its in canonical_to_items.items():
                if canonical_key in can_k or can_k in canonical_key:
                    for it in its:
                        it["Quantity"] = float(round(qty, 4))
                        it["Justification"] = just
                    matched = True
                    if DIAGNOSTIC:
                        print("Fuzzy matched", canonical_key, "->", can_k)
                    break
            if not matched and DIAGNOSTIC:
                print("No template item matched for computed key:", canonical_key)

    # As a final fallback, if some items remain zero but agg contains numeric totals, try to map them
    for key, payload in agg.items():
        total_qty = sum(payload.get("units", {}).values())
        if total_qty <= 0:
            continue
        # try to find template items whose canonical contains key
        for can_k, its in canonical_to_items.items():
            if key in can_k or can_k in key:
                for it in its:
                    if it["Quantity"] == 0:
                        it["Quantity"] = float(round(total_qty, 4))
                        it["Justification"] = "Fallback mapped from parsed totals"
                        if DIAGNOSTIC:
                            print("Fallback mapped", key, "->", can_k, total_qty)

    # Rate lookup and amount calc
    for it in items:
        try:
            rate = get_rate_from_library(it["Canonical"], it["Description"], it["Unit"], location=location)
        except Exception:
            rate = None
        if rate is None:
            rate = DEFAULT_RATE
        it["Rate"] = float(rate)
        it["Amount"] = round(it["Quantity"] * it["Rate"], 2)

    df = pd.DataFrame(items)
    return df

def export_besmm4_using_template(populated_df: pd.DataFrame, output_path: str, template_path: str = None):
    if template_path is None:
        template_path = TEMPLATE_PATH

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"BESMM4 template not found at {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    # find header row and columns
    header_row = None
    item_code_col = None
    qty_col = None
    rate_col = None
    amount_col = None

    search_limit = min(60, ws.max_row)
    for r in range(1, search_limit + 1):
        values = [ (ws.cell(row=r, column=c).value if ws.cell(row=r, column=c).value is not None else "") for c in range(1, ws.max_column+1) ]
        joined = "|".join([str(v).strip().lower() for v in values if v is not None])
        if "item code" in joined or "itemcode" in joined:
            header_row = r
            for c, val in enumerate(values, start=1):
                if not val:
                    continue
                v = str(val).strip().lower()
                if v in ("item code", "itemcode", "item_code"):
                    item_code_col = c
                if "quantity" in v and qty_col is None:
                    qty_col = c
                if "rate" in v and rate_col is None:
                    rate_col = c
                if "amount" in v and amount_col is None:
                    amount_col = c
            break

    if header_row is None or item_code_col is None or qty_col is None:
        item_code_col = item_code_col or 1
        qty_col = qty_col or 4
        rate_col = rate_col or 5
        amount_col = amount_col or 6
        header_row = header_row or 1

    # build index item_code -> row
    item_to_row = {}
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=item_code_col)
        if cell.value is None:
            continue
        code = str(cell.value).strip()
        if code == "":
            continue
        if code not in item_to_row:
            item_to_row[code] = r

    # write values
    for _, row in populated_df.iterrows():
        item_code = str(row["ItemCode"]).strip()
        if item_code in item_to_row:
            r = item_to_row[item_code]
            qcell = ws.cell(row=r, column=qty_col)
            qcell.value = float(row["Quantity"] or 0.0)
            try:
                if not qcell.number_format or qcell.number_format == 'General':
                    qcell.number_format = "#,##0.##"
            except Exception:
                pass

            rcell = ws.cell(row=r, column=rate_col)
            rcell.value = float(row["Rate"] or 0.0)
            try:
                if not rcell.number_format or rcell.number_format == 'General':
                    rcell.number_format = '"₦"#,##0.00'
            except Exception:
                pass

            acell = ws.cell(row=r, column=amount_col)
            acell.value = float(row["Amount"] or 0.0)
            try:
                if not acell.number_format or acell.number_format == 'General':
                    acell.number_format = '"₦"#,##0.00'
            except Exception:
                pass
        else:
            # not found in template; skip
            if DIAGNOSTIC:
                print("Template missing item code:", item_code)

    # audit sheet
    audit_name = "_audit"
    if audit_name in wb.sheetnames:
        wb.remove(wb[audit_name])
    audit = wb.create_sheet(audit_name)
    audit.append(["ItemCode", "Description", "Unit", "Quantity", "Rate", "Amount", "Justification"])
    for _, row in populated_df.iterrows():
        audit.append([
            row["ItemCode"],
            row["Description"],
            row["Unit"],
            float(row["Quantity"] or 0.0),
            float(row["Rate"] or 0.0),
            float(row["Amount"] or 0.0),
            row.get("Justification", "")
        ])
    audit.sheet_state = "hidden"

    wb.save(output_path)
    return output_path

# public
def generate_besmm4_boq(parsed_entries: list, output_path: str, location: str = "Kaduna", template_path: str = None, diagnostic: bool = False):
    """
    Entry point.
    """
    global DIAGNOSTIC
    DIAGNOSTIC = bool(diagnostic)

    if not isinstance(parsed_entries, list):
        raise ValueError("parsed_entries must be a list of dicts")

    if template_path is None:
        template_path = TEMPLATE_PATH

    populated = populate_besmm4_from_parsed(parsed_entries, location=location, include_empty=True)
    out = export_besmm4_using_template(populated, output_path, template_path=template_path)
    if DIAGNOSTIC:
        print("Generated:", out)
    return out

if __name__ == "__main__":
    print("This module provides generate_besmm4_boq(parsed_entries, output_path, ...)")



